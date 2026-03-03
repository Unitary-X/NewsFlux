"""
Report generation service — PDF (with NewsFlux watermark) and Excel exports.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen.canvas import Canvas


# ────────────────────────────────────
# WATERMARK CANVAS
# ────────────────────────────────────

class WatermarkCanvas(Canvas):
    """Custom canvas that draws a 'NewsFlux' watermark on every page."""

    def showPage(self):
        self._draw_watermark()
        super().showPage()

    def _draw_watermark(self):
        self.saveState()
        self.setFont("Helvetica-Bold", 54)
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.06))
        w, h = A4
        # Diagonal watermark centred on the page
        self.translate(w / 2, h / 2)
        self.rotate(45)
        self.drawCentredString(0, 0, "NewsFlux")
        # Second copy offset for visual coverage
        self.drawCentredString(0, 160, "NewsFlux")
        self.drawCentredString(0, -160, "NewsFlux")
        self.restoreState()


def _pdf_doc(buf, title):
    """Return a SimpleDocTemplate wired to our watermark canvas."""
    return SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=title,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
    )


# Shared styles
_styles = getSampleStyleSheet()
_title_style = ParagraphStyle("RTitle", parent=_styles["Heading1"], fontSize=16, spaceAfter=6)
_sub_style = ParagraphStyle("RSub", parent=_styles["Normal"], fontSize=9, textColor=colors.grey)
_section_style = ParagraphStyle("RSec", parent=_styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=4)

_header_bg = colors.HexColor("#1e3a5f")
_header_fg = colors.white
_alt_row = colors.HexColor("#f8fafc")


def _make_table(headers, rows, col_widths=None):
    """Build a styled Platypus Table."""
    data = [headers] + rows
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), _header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), _header_fg),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _alt_row]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(style))
    return t


def _fmt(v):
    """Format a number for display."""
    if isinstance(v, float):
        return f"₹{v:,.2f}" if v >= 0 else f"-₹{abs(v):,.2f}"
    return str(v)


# ────────────────────────────────────
# EXCEL HELPERS
# ────────────────────────────────────

_xl_header_font = Font(bold=True, color="FFFFFF", size=10)
_xl_header_fill = PatternFill("solid", fgColor="1e3a5f")
_xl_header_align = Alignment(horizontal="center", vertical="center")
_xl_border = Border(
    left=Side(style="thin", color="cbd5e1"),
    right=Side(style="thin", color="cbd5e1"),
    top=Side(style="thin", color="cbd5e1"),
    bottom=Side(style="thin", color="cbd5e1"),
)


def _xl_write_header(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _xl_header_font
        cell.fill = _xl_header_fill
        cell.alignment = _xl_header_align
        cell.border = _xl_border


def _xl_write_rows(ws, rows, start_row=2):
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _xl_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="f8fafc")


def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ────────────────────────────────────
# PROFIT & LOSS
# ────────────────────────────────────

def profit_loss_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, "Profit & Loss Report")
    elems = []
    elems.append(Paragraph("Profit & Loss Report", _title_style))
    elems.append(Paragraph(f"{_month_name(data['month'])} {data['year']}", _sub_style))
    elems.append(Spacer(1, 10))

    # KPI summary
    kpi = [
        ["Metric", "Amount"],
        ["Total Revenue", _fmt(data["revenue"]["total"])],
        ["  Collected", _fmt(data["revenue"]["collected"])],
        ["  Pending", _fmt(data["revenue"]["pending"])],
        ["Newspaper Purchase Cost", _fmt(data["expenses"]["purchase_cost"])],
        ["Worker Salaries", _fmt(data["expenses"]["salary"])],
        ["Total Expenses", _fmt(data["expenses"]["total"])],
        ["Net Profit", _fmt(data["net_profit"])],
    ]
    elems.append(_make_table(kpi[0], kpi[1:], col_widths=[120 * mm, 50 * mm]))
    elems.append(Spacer(1, 10))

    # Stock summary
    elems.append(Paragraph("Stock Summary", _section_style))
    stock = [
        ["Taken", "Returned", "Sold"],
        [data["stock"]["taken"], data["stock"]["returned"], data["stock"]["sold"]],
    ]
    elems.append(_make_table(stock[0], stock[1:]))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph(f"Invoices: {data['invoices_count']}  |  Salary Records: {data['salaries_count']}", _sub_style))

    doc.build(elems, canvasmaker=WatermarkCanvas)
    return buf.getvalue()


def profit_loss_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"

    ws.cell(row=1, column=1, value="Profit & Loss Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{_month_name(data['month'])} {data['year']}")

    headers = ["Metric", "Amount"]
    _xl_write_header(ws, headers, row=4)
    rows = [
        ("Total Revenue", data["revenue"]["total"]),
        ("  Collected", data["revenue"]["collected"]),
        ("  Pending", data["revenue"]["pending"]),
        ("Newspaper Purchase Cost", data["expenses"]["purchase_cost"]),
        ("Worker Salaries", data["expenses"]["salary"]),
        ("Total Expenses", data["expenses"]["total"]),
        ("Net Profit", data["net_profit"]),
    ]
    _xl_write_rows(ws, rows, start_row=5)

    # Stock summary
    r = 5 + len(rows) + 1
    ws.cell(row=r, column=1, value="Stock Summary").font = Font(bold=True, size=11)
    _xl_write_header(ws, ["Taken", "Returned", "Sold"], row=r + 1)
    _xl_write_rows(ws, [(data["stock"]["taken"], data["stock"]["returned"], data["stock"]["sold"])], start_row=r + 2)
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────
# STOCK RECONCILIATION
# ────────────────────────────────────

def stock_recon_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, "Stock Reconciliation")
    elems = []
    elems.append(Paragraph("Stock Reconciliation Report", _title_style))
    elems.append(Paragraph(f"Date: {data['date']}", _sub_style))
    elems.append(Spacer(1, 10))

    # Summary
    s = data["summary"]
    summary = [
        ["Matched", "Surplus", "Deficit", "Total Expected", "Total Sold", "Discrepancy"],
        [s["newspapers_matched"], s["newspapers_surplus"], s["newspapers_deficit"],
         s["total_expected"], s["total_sold"], s["total_discrepancy"]],
    ]
    elems.append(_make_table(summary[0], summary[1:]))
    elems.append(Spacer(1, 10))

    # Detail table
    elems.append(Paragraph("Newspaper Detail", _section_style))
    headers = ["Newspaper", "Expected", "Taken", "Returned", "Sold", "Discrepancy", "Status"]
    rows = [[n["newspaper_name"], n["expected"], n["taken"], n["returned"],
             n["sold"], n["discrepancy"], n["status"].title()] for n in data["newspapers"]]
    elems.append(_make_table(headers, rows))

    doc.build(elems, canvasmaker=WatermarkCanvas)
    return buf.getvalue()


def stock_recon_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Reconciliation"

    ws.cell(row=1, column=1, value="Stock Reconciliation Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Date: {data['date']}")

    headers = ["Newspaper", "Expected", "Taken", "Returned", "Sold", "Discrepancy", "Status"]
    _xl_write_header(ws, headers, row=4)
    rows = [(n["newspaper_name"], n["expected"], n["taken"], n["returned"],
             n["sold"], n["discrepancy"], n["status"].title()) for n in data["newspapers"]]
    _xl_write_rows(ws, rows, start_row=5)
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────
# WORKER PERFORMANCE
# ────────────────────────────────────

def worker_perf_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, "Worker Performance")
    elems = []
    elems.append(Paragraph("Worker Performance Report", _title_style))
    elems.append(Paragraph(f"{_month_name(data['month'])} {data['year']}", _sub_style))
    elems.append(Spacer(1, 10))

    headers = ["#", "Worker", "Assignments", "Delivered", "Missed", "Rate %", "Salary (₹)", "Status"]
    rows = []
    for i, w in enumerate(data["workers"], 1):
        rows.append([
            i, w["worker_name"], w["assignments"], w["delivered"],
            w["missed"], f"{w['delivery_rate']}%",
            f"{w['salary_amount']:,.2f}", w["salary_status"],
        ])
    elems.append(_make_table(headers, rows))

    doc.build(elems, canvasmaker=WatermarkCanvas)
    return buf.getvalue()


def worker_perf_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worker Performance"

    ws.cell(row=1, column=1, value="Worker Performance Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{_month_name(data['month'])} {data['year']}")

    headers = ["#", "Worker", "Assignments", "Delivered", "Missed", "Rate %", "Salary", "Status"]
    _xl_write_header(ws, headers, row=4)
    rows = []
    for i, w in enumerate(data["workers"], 1):
        rows.append((i, w["worker_name"], w["assignments"], w["delivered"],
                      w["missed"], w["delivery_rate"], w["salary_amount"], w["salary_status"]))
    _xl_write_rows(ws, rows, start_row=5)
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────
# SUMMARY
# ────────────────────────────────────

def summary_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, "Report Summary")
    elems = []
    elems.append(Paragraph("Report Summary", _title_style))
    elems.append(Paragraph(
        f"Period: {data['period'].title()}  |  {data['start_date']} to {data['end_date']}", _sub_style
    ))
    elems.append(Spacer(1, 10))

    # Overview
    elems.append(Paragraph("Overview", _section_style))
    kpi = [
        ["Metric", "Value"],
        ["Revenue", _fmt(data["revenue"])],
        ["Stock Taken", str(data["stock"]["taken"])],
        ["Stock Returned", str(data["stock"]["returned"])],
        ["Stock Sold", str(data["stock"]["sold"])],
        ["Total Deliveries", str(data["deliveries"]["total"])],
        ["Delivered", str(data["deliveries"]["delivered"])],
        ["Missed", str(data["deliveries"]["missed"])],
    ]
    elems.append(_make_table(kpi[0], kpi[1:], col_widths=[100 * mm, 70 * mm]))
    elems.append(Spacer(1, 10))

    # Daily breakdown
    if data.get("daily_breakdown"):
        elems.append(Paragraph("Daily Breakdown", _section_style))
        headers = ["Date", "Taken", "Returned", "Sold", "Revenue (₹)"]
        rows = [[d["date"], d["taken"], d["returned"], d["sold"], f"{d['revenue']:,.2f}"]
                for d in data["daily_breakdown"]]
        elems.append(_make_table(headers, rows))

    doc.build(elems, canvasmaker=WatermarkCanvas)
    return buf.getvalue()


def summary_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="Report Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {data['period'].title()}  |  {data['start_date']} to {data['end_date']}")

    # Overview
    _xl_write_header(ws, ["Metric", "Value"], row=4)
    overview = [
        ("Revenue", data["revenue"]),
        ("Stock Taken", data["stock"]["taken"]),
        ("Stock Returned", data["stock"]["returned"]),
        ("Stock Sold", data["stock"]["sold"]),
        ("Total Deliveries", data["deliveries"]["total"]),
        ("Delivered", data["deliveries"]["delivered"]),
        ("Missed", data["deliveries"]["missed"]),
    ]
    _xl_write_rows(ws, overview, start_row=5)

    # Daily breakdown
    if data.get("daily_breakdown"):
        r = 5 + len(overview) + 1
        ws.cell(row=r, column=1, value="Daily Breakdown").font = Font(bold=True, size=11)
        _xl_write_header(ws, ["Date", "Taken", "Returned", "Sold", "Revenue"], row=r + 1)
        rows = [(d["date"], d["taken"], d["returned"], d["sold"], d["revenue"])
                for d in data["daily_breakdown"]]
        _xl_write_rows(ws, rows, start_row=r + 2)
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────
# UTILS
# ────────────────────────────────────

def _month_name(m: int) -> str:
    return date(2000, m, 1).strftime("%B")
