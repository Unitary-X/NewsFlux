"""
Excel Export Service — Generates .xlsx reports per agency for Google Drive backup.
Three report types: Daily, Monthly, Yearly.
"""
import io
import calendar
from datetime import date, datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, Date as SaDate

from app.models.models import (
    Newspaper, DailyStock, Customer, CustomerSubscription,
    WorkerAssignment, Invoice, User
)

# ── Styling helpers ──────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _to_float(val):
    if isinstance(val, Decimal):
        return float(val)
    return val or 0


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Daily Report ─────────────────────────────────────────────────

def generate_daily_stock_excel(db: Session, tenant_id, target_date: date) -> bytes:
    """Daily stock report: per-newspaper taken/returned/sold/revenue."""
    newspapers = db.query(Newspaper).filter(Newspaper.tenant_id == tenant_id).all()
    stocks = db.query(DailyStock).filter(
        DailyStock.tenant_id == tenant_id,
        DailyStock.date == target_date,
    ).all()
    stock_map = {s.newspaper_id: s for s in stocks}

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Stock"

    headers = ["Newspaper", "Taken", "Returned", "Sold", "Base Price", "Revenue"]
    ws.append(headers)
    _style_header(ws, len(headers))

    total_revenue = 0
    for np in newspapers:
        s = stock_map.get(np.id)
        taken = s.taken if s else 0
        returned = s.returned if s else 0
        sold = taken - returned
        price = _to_float(np.base_price)
        revenue = sold * price
        total_revenue += revenue
        ws.append([np.name, taken, returned, sold, price, revenue])

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", total_revenue])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def generate_daily_deliveries_excel(db: Session, tenant_id, target_date: date) -> bytes:
    """Daily deliveries report: per-worker customer assignments."""
    assignments = db.query(WorkerAssignment).filter(
        WorkerAssignment.tenant_id == tenant_id,
    ).order_by(WorkerAssignment.worker_id, WorkerAssignment.route_order).all()

    workers = db.query(User).filter(
        User.tenant_id == tenant_id, User.role == "worker"
    ).all()
    worker_map = {w.id: w.username for w in workers}

    customers = db.query(Customer).filter(Customer.tenant_id == tenant_id).all()
    customer_map = {c.id: c.name for c in customers}

    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries"

    headers = ["Worker", "Route Order", "Customer", "Customer Phone"]
    ws.append(headers)
    _style_header(ws, len(headers))

    cust_detail_map = {c.id: c for c in customers}
    for a in assignments:
        worker_name = worker_map.get(a.worker_id, "Unknown")
        cust = cust_detail_map.get(a.customer_id)
        ws.append([
            worker_name,
            a.route_order,
            cust.name if cust else "Unknown",
            cust.phone if cust else "",
        ])

    _auto_width(ws)
    return _workbook_to_bytes(wb)


# ── Monthly Report ───────────────────────────────────────────────

def generate_monthly_revenue_excel(db: Session, tenant_id, month: int, year: int) -> bytes:
    """Monthly revenue report: daily revenue per newspaper over the month."""
    newspapers = db.query(Newspaper).filter(Newspaper.tenant_id == tenant_id).all()
    days_in_month = calendar.monthrange(year, month)[1]

    stocks = db.query(DailyStock).filter(
        DailyStock.tenant_id == tenant_id,
        func.extract('month', DailyStock.date) == month,
        func.extract('year', DailyStock.date) == year,
    ).all()

    # Build lookup: (newspaper_id, day) -> stock
    stock_lookup = {}
    for s in stocks:
        stock_lookup[(s.newspaper_id, s.date.day)] = s

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    headers = ["Day"] + [np.name for np in newspapers] + ["Daily Total"]
    ws.append(headers)
    _style_header(ws, len(headers))

    grand_total = 0
    for day in range(1, days_in_month + 1):
        row = [day]
        daily_total = 0
        for np in newspapers:
            s = stock_lookup.get((np.id, day))
            if s:
                sold = (s.taken or 0) - (s.returned or 0)
                rev = sold * _to_float(np.base_price)
            else:
                rev = 0
            daily_total += rev
            row.append(rev)
        row.append(daily_total)
        grand_total += daily_total
        ws.append(row)

    ws.append([])
    total_row = ["TOTAL"] + [""] * len(newspapers) + [grand_total]
    ws.append(total_row)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def generate_monthly_subscriptions_excel(db: Session, tenant_id, month: int, year: int) -> bytes:
    """Monthly subscription summary: customer × newspaper matrix."""
    subs = db.query(CustomerSubscription).filter(
        CustomerSubscription.tenant_id == tenant_id,
    ).all()
    customers = db.query(Customer).filter(Customer.tenant_id == tenant_id).all()
    newspapers = db.query(Newspaper).filter(Newspaper.tenant_id == tenant_id).all()

    cust_map = {c.id: c.name for c in customers}
    np_map = {n.id: n.name for n in newspapers}

    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions"

    headers = ["Customer"] + [np_map.get(n.id, "?") for n in newspapers] + ["Status Count"]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Group subs by customer
    cust_subs = {}
    for s in subs:
        cust_subs.setdefault(s.customer_id, []).append(s)

    for cust in customers:
        row = [cust.name]
        active_count = 0
        my_subs = {s.newspaper_id: s for s in cust_subs.get(cust.id, [])}
        for np in newspapers:
            s = my_subs.get(np.id)
            if s and s.status == 1:
                row.append(f"Active (×{s.quantity})")
                active_count += 1
            elif s:
                row.append("Paused")
            else:
                row.append("-")
        row.append(active_count)
        ws.append(row)

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def generate_monthly_invoices_excel(db: Session, tenant_id, month: int, year: int) -> bytes:
    """Monthly invoice report: all invoices for the period."""
    invoices = db.query(Invoice).filter(
        Invoice.tenant_id == tenant_id,
        Invoice.month == month,
        Invoice.year == year,
    ).all()
    customers = db.query(Customer).filter(Customer.tenant_id == tenant_id).all()
    cust_map = {c.id: c.name for c in customers}

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = ["Customer", "Total Amount", "Delivery Fee", "Status"]
    ws.append(headers)
    _style_header(ws, len(headers))

    total_amt = 0
    total_fee = 0
    for inv in invoices:
        amt = _to_float(inv.total_amount)
        fee = _to_float(inv.delivery_fee)
        total_amt += amt
        total_fee += fee
        ws.append([
            cust_map.get(inv.customer_id, "Unknown"),
            amt,
            fee,
            inv.status.capitalize(),
        ])

    ws.append([])
    ws.append(["TOTAL", total_amt, total_fee, ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _auto_width(ws)
    return _workbook_to_bytes(wb)


# ── Yearly Report ────────────────────────────────────────────────

def generate_yearly_report_excel(db: Session, tenant_id, year: int) -> bytes:
    """Yearly report: monthly revenue trend + totals."""
    newspapers = db.query(Newspaper).filter(Newspaper.tenant_id == tenant_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Revenue"

    headers = ["Month"] + [np.name for np in newspapers] + ["Monthly Total"]
    ws.append(headers)
    _style_header(ws, len(headers))

    grand_total = 0
    for month in range(1, 13):
        stocks = db.query(DailyStock).filter(
            DailyStock.tenant_id == tenant_id,
            func.extract('month', DailyStock.date) == month,
            func.extract('year', DailyStock.date) == year,
        ).all()

        np_revenue = {}
        for s in stocks:
            sold = (s.taken or 0) - (s.returned or 0)
            np_obj = next((n for n in newspapers if n.id == s.newspaper_id), None)
            price = _to_float(np_obj.base_price) if np_obj else 0
            np_revenue[s.newspaper_id] = np_revenue.get(s.newspaper_id, 0) + (sold * price)

        month_name = calendar.month_abbr[month]
        row = [month_name]
        monthly_total = 0
        for np in newspapers:
            rev = np_revenue.get(np.id, 0)
            monthly_total += rev
            row.append(rev)
        row.append(monthly_total)
        grand_total += monthly_total
        ws.append(row)

    ws.append([])
    ws.append(["TOTAL"] + [""] * len(newspapers) + [grand_total])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Sheet 2: Customer growth
    ws2 = wb.create_sheet("Customer Growth")
    ws2.append(["Month", "Total Customers"])
    _style_header(ws2, 2)
    customers = db.query(Customer).filter(Customer.tenant_id == tenant_id).all()
    for month in range(1, 13):
        count = sum(1 for c in customers if c.id)  # All customers (no created_at tracking)
        ws2.append([calendar.month_abbr[month], count])
    _auto_width(ws2)

    # Sheet 3: Invoice summary
    ws3 = wb.create_sheet("Invoice Summary")
    ws3.append(["Month", "Total Invoices", "Pending", "Paid", "Total Amount"])
    _style_header(ws3, 5)
    for month in range(1, 13):
        invs = db.query(Invoice).filter(
            Invoice.tenant_id == tenant_id,
            Invoice.month == month,
            Invoice.year == year,
        ).all()
        pending = sum(1 for i in invs if i.status == "pending")
        paid = sum(1 for i in invs if i.status == "paid")
        total = sum(_to_float(i.total_amount) for i in invs)
        ws3.append([calendar.month_abbr[month], len(invs), pending, paid, total])
    _auto_width(ws3)

    _auto_width(ws)
    return _workbook_to_bytes(wb)
