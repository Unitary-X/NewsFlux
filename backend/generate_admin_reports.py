"""Generate monthly admin Excel reports from JSON backups.

This script scans JSON files recursively, extracts record dates, and writes
organized Excel reports in year/month folders.
"""

from __future__ import annotations

import calendar
import json
import logging
import math
import re
from collections import defaultdict
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ==========================
# Configurable entry points
# ==========================
INPUT_DIR = Path("./backups")
OUTPUT_DIR = Path("./Admin_Reports")
DATE_FIELD = "date"
CATEGORY_FIELD = "category"

# Fallback date fields (checked after DATE_FIELD).
FALLBACK_DATE_FIELDS: Sequence[str] = ("created_at", "timestamp")

ERROR_LOG_NAME = "errors.log"

SUMMARY_SHEET = "Summary"
DETAIL_SHEET = "Detailed Records"
CATEGORY_SHEET = "Category Breakdown"

HASH_FIELD = "__record_hash"

HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_FILL_ALT = PatternFill(fill_type="solid", start_color="F2F6FC", end_color="F2F6FC")


def build_loggers(output_dir: Path) -> Tuple[logging.Logger, logging.Logger]:
    """Create application and error-only loggers."""
    output_dir.mkdir(parents=True, exist_ok=True)

    app_logger = logging.getLogger("admin_reports")
    app_logger.setLevel(logging.INFO)
    if not app_logger.handlers:
        stream = logging.StreamHandler()
        stream.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        app_logger.addHandler(stream)

    error_logger = logging.getLogger("admin_reports.errors")
    error_logger.setLevel(logging.WARNING)
    if not error_logger.handlers:
        fh = logging.FileHandler(output_dir / ERROR_LOG_NAME, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        error_logger.addHandler(fh)

    return app_logger, error_logger


def coerce_records(payload: Any) -> List[Dict[str, Any]]:
    """Normalize supported JSON payload shapes into a list of dict records."""
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]

    if isinstance(payload, dict):
        for key in ("records", "data", "items", "transactions", "results"):
            nested = payload.get(key)
            if isinstance(nested, list):
                return [item for item in nested if isinstance(item, dict)]
        return [payload]

    return []


def try_parse_datetime(value: Any) -> Optional[datetime]:
    """Parse multiple date/time representations safely."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, (int, float)):
        if not math.isfinite(value):
            return None

        # Heuristic for epoch milliseconds vs epoch seconds.
        ts = float(value)
        if abs(ts) > 1e12:
            ts = ts / 1000.0
        try:
            return datetime.fromtimestamp(ts)
        except (OverflowError, OSError, ValueError):
            return None

    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    # Numeric strings as epoch timestamps.
    if re.fullmatch(r"[-+]?\d+(\.\d+)?", text):
        try:
            numeric = float(text)
            return try_parse_datetime(numeric)
        except ValueError:
            return None

    # ISO-8601 variants (including trailing Z).
    iso_candidate = text.replace("Z", "+00:00") if text.endswith("Z") else text
    try:
        return datetime.fromisoformat(iso_candidate)
    except ValueError:
        pass

    known_formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%d %b %Y",
        "%d %B %Y",
    )

    for fmt in known_formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    # Last chance: extract a leading YYYY-MM-DD style token.
    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        try:
            year, month, day = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return datetime(year=year, month=month, day=day)
        except ValueError:
            return None

    return None


def extract_record_datetime(record: Dict[str, Any], date_field: str) -> Optional[datetime]:
    """Read and parse record date from configured and fallback fields."""
    candidates = [date_field, *FALLBACK_DATE_FIELDS]
    for field in candidates:
        if field not in record:
            continue
        parsed = try_parse_datetime(record.get(field))
        if parsed:
            return parsed
    return None


def canonical_hash(record: Dict[str, Any]) -> str:
    """Compute stable hash for de-duplication."""
    serializable = {k: v for k, v in record.items() if k != HASH_FIELD}
    payload = json.dumps(serializable, ensure_ascii=False, sort_keys=True, default=str)
    return sha256(payload.encode("utf-8")).hexdigest()


def normalize_cell_value(value: Any) -> Any:
    """Convert unsupported values to printable text for Excel cells."""
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return json.dumps(value, ensure_ascii=False, default=str)


def style_header_row(ws) -> None:
    """Apply header styling and freeze pane."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def apply_alternating_rows(ws) -> None:
    """Apply zebra striping for data rows."""
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = ROW_FILL_ALT


def autofit_columns(ws) -> None:
    """Set width based on current content length."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def clear_sheet(ws) -> None:
    """Remove all content from a sheet."""
    if ws.max_row > 1:
        ws.delete_rows(1, ws.max_row)


def get_or_create_sheet(wb: Workbook, title: str):
    """Get sheet by title, creating it when absent."""
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def read_existing_hashes_and_headers(ws) -> Tuple[set, List[str]]:
    """Read existing dedupe hashes and header columns from detail sheet."""
    if ws.max_row < 1 or ws.max_column < 1 or ws.cell(1, 1).value is None:
        return set(), []

    headers = [
        (ws.cell(1, c).value if ws.cell(1, c).value is not None else "")
        for c in range(1, ws.max_column + 1)
    ]
    headers = [str(h) for h in headers if str(h)]

    if not headers:
        return set(), []

    existing_hashes = set()
    if HASH_FIELD in headers:
        hash_col = headers.index(HASH_FIELD) + 1
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=hash_col).value
            if value:
                existing_hashes.add(str(value))

    return existing_hashes, headers


def gather_columns(existing_headers: List[str], records: List[Dict[str, Any]]) -> List[str]:
    """Merge existing columns with new record keys, preserving order."""
    columns = [h for h in existing_headers if h]

    for record in records:
        for key in record.keys():
            if key not in columns and key != HASH_FIELD:
                columns.append(key)

    if HASH_FIELD not in columns:
        columns.append(HASH_FIELD)

    return columns


def parse_float(value: Any) -> Optional[float]:
    """Best effort numeric conversion."""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return float(value)

    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None

    return None


def read_detail_records(ws, columns: List[str]) -> List[Dict[str, Any]]:
    """Read all detail rows into dictionaries."""
    records: List[Dict[str, Any]] = []
    if ws.max_row < 2:
        return records

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns), values_only=True):
        row_data = {columns[idx]: row[idx] for idx in range(len(columns))}
        if all(value is None for value in row_data.values()):
            continue
        records.append(row_data)
    return records


def write_summary_sheet(ws, year: int, month: int, records: List[Dict[str, Any]], category_field: str) -> None:
    """Create summary sheet content for one month."""
    clear_sheet(ws)

    month_name = calendar.month_name[month]
    ws["A1"] = f"Admin Report - {month_name} {year}"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "Month"
    ws["B3"] = month_name
    ws["A4"] = "Year"
    ws["B4"] = year
    ws["A5"] = "Total Records"
    ws["B5"] = len(records)

    row = 7

    # Category breakdown in summary if available.
    categories = defaultdict(int)
    for record in records:
        category = record.get(category_field)
        if category is not None and str(category).strip():
            categories[str(category).strip()] += 1

    if categories:
        ws.cell(row=row, column=1, value="Category Breakdown")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Count")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1
        for category, count in sorted(categories.items()):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
            row += 1
        row += 1

    # Numeric key totals.
    totals: Dict[str, float] = defaultdict(float)
    for record in records:
        for key, value in record.items():
            if key == HASH_FIELD:
                continue
            numeric = parse_float(value)
            if numeric is None:
                continue
            totals[key] += numeric

    if totals:
        ws.cell(row=row, column=1, value="Key Totals")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Field")
        ws.cell(row=row, column=2, value="Total")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1
        for key, total in sorted(totals.items()):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=round(total, 2))
            row += 1

    autofit_columns(ws)


def write_category_sheet(ws, records: List[Dict[str, Any]], columns: List[str], category_field: str) -> None:
    """Create category grouping sheet with subtotals."""
    clear_sheet(ws)

    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        category_value = record.get(category_field)
        if category_value is None or not str(category_value).strip():
            continue
        grouped[str(category_value).strip()].append(record)

    if not grouped:
        ws["A1"] = "No category data available for this month."
        autofit_columns(ws)
        return

    ws.append(["Category", "Record Count"])
    numeric_columns = [col for col in columns if col != HASH_FIELD]

    numeric_totals_columns: List[str] = []
    for col in numeric_columns:
        has_numeric = any(parse_float(rec.get(col)) is not None for rec in records)
        if has_numeric:
            numeric_totals_columns.append(col)
            ws.cell(row=1, column=ws.max_column + 1, value=f"Total {col}")

    style_header_row(ws)

    for category, recs in sorted(grouped.items()):
        row_values: List[Any] = [category, len(recs)]
        for col in numeric_totals_columns:
            subtotal = 0.0
            for rec in recs:
                val = parse_float(rec.get(col))
                if val is not None:
                    subtotal += val
            row_values.append(round(subtotal, 2))
        ws.append(row_values)

    ws.auto_filter.ref = ws.dimensions
    apply_alternating_rows(ws)
    autofit_columns(ws)


def write_or_update_month_report(
    report_path: Path,
    year: int,
    month: int,
    incoming_records: List[Dict[str, Any]],
    category_field: str,
) -> Tuple[int, int]:
    """Write monthly report and append de-duplicated records.

    Returns:
        Tuple[records_written, duplicates_skipped]
    """
    if report_path.exists():
        wb = load_workbook(report_path)
    else:
        wb = Workbook()

    if wb.sheetnames == ["Sheet"]:
        del wb["Sheet"]

    detail_ws = get_or_create_sheet(wb, DETAIL_SHEET)
    summary_ws = get_or_create_sheet(wb, SUMMARY_SHEET)

    existing_hashes, existing_headers = read_existing_hashes_and_headers(detail_ws)

    deduped_records: List[Dict[str, Any]] = []
    duplicates = 0
    for record in incoming_records:
        rec = dict(record)
        rec[HASH_FIELD] = canonical_hash(rec)
        if rec[HASH_FIELD] in existing_hashes:
            duplicates += 1
            continue
        existing_hashes.add(rec[HASH_FIELD])
        deduped_records.append(rec)

    columns = gather_columns(existing_headers, deduped_records)

    # Initialize header only once when sheet is empty.
    if detail_ws.max_row < 1 or detail_ws.cell(1, 1).value is None:
        detail_ws.append(columns)
    else:
        # Expand existing header if new columns appear.
        current_headers = [detail_ws.cell(1, c).value for c in range(1, detail_ws.max_column + 1)]
        current_headers = [str(h) for h in current_headers if h is not None]
        if current_headers != columns:
            # Re-write full sheet preserving existing data if schema changed.
            existing_rows = read_detail_records(detail_ws, current_headers)
            clear_sheet(detail_ws)
            detail_ws.append(columns)
            for row_dict in existing_rows:
                row_values = [normalize_cell_value(row_dict.get(col)) for col in columns]
                detail_ws.append(row_values)

    for record in deduped_records:
        row_values = [normalize_cell_value(record.get(col)) for col in columns]
        detail_ws.append(row_values)

    # Formatting and usability for details.
    style_header_row(detail_ws)
    detail_ws.auto_filter.ref = detail_ws.dimensions
    apply_alternating_rows(detail_ws)
    autofit_columns(detail_ws)

    if HASH_FIELD in columns:
        hidden_col_idx = columns.index(HASH_FIELD) + 1
        detail_ws.column_dimensions[get_column_letter(hidden_col_idx)].hidden = True

    # Build summaries from all current rows in this workbook.
    all_records = read_detail_records(detail_ws, columns)
    write_summary_sheet(summary_ws, year, month, all_records, category_field)

    has_category_values = any(
        rec.get(category_field) is not None and str(rec.get(category_field)).strip()
        for rec in all_records
    )
    if has_category_values:
        category_ws = get_or_create_sheet(wb, CATEGORY_SHEET)
        write_category_sheet(category_ws, all_records, columns, category_field)
    else:
        if CATEGORY_SHEET in wb.sheetnames:
            del wb[CATEGORY_SHEET]

    # Keep predictable sheet order.
    desired_order = [SUMMARY_SHEET, DETAIL_SHEET]
    if CATEGORY_SHEET in wb.sheetnames:
        desired_order.append(CATEGORY_SHEET)
    wb._sheets.sort(key=lambda s: desired_order.index(s.title) if s.title in desired_order else 99)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)

    return len(deduped_records), duplicates


def process_json_backups(
    input_dir: Path,
    output_dir: Path,
    date_field: str,
    category_field: str,
) -> Dict[str, int]:
    """Process JSON backups and produce monthly reports."""
    app_logger, error_logger = build_loggers(output_dir)

    stats = {
        "files_processed": 0,
        "records_seen": 0,
        "records_written": 0,
        "duplicates_skipped": 0,
        "invalid_date_skipped": 0,
        "malformed_files": 0,
        "folders_created": 0,
    }

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    grouped_records: Dict[Tuple[int, int], List[Dict[str, Any]]] = defaultdict(list)

    json_files = sorted(input_dir.rglob("*.json"))
    for json_file in json_files:
        stats["files_processed"] += 1
        try:
            with json_file.open("r", encoding="utf-8") as fh:
                payload = json.load(fh)
        except (OSError, json.JSONDecodeError) as exc:
            stats["malformed_files"] += 1
            error_logger.warning("Malformed or unreadable JSON file skipped: %s (%s)", json_file, exc)
            continue

        records = coerce_records(payload)
        for record in records:
            stats["records_seen"] += 1
            record_dt = extract_record_datetime(record, date_field)
            if not record_dt:
                stats["invalid_date_skipped"] += 1
                error_logger.warning(
                    "Record skipped due to invalid/missing date in %s: %s",
                    json_file,
                    json.dumps(record, ensure_ascii=False, default=str),
                )
                continue

            grouped_records[(record_dt.year, record_dt.month)].append(record)

    created_dirs: set = set()
    for (year, month) in sorted(grouped_records.keys()):
        month_name = calendar.month_name[month]
        month_dir = output_dir / str(year) / month_name
        if not month_dir.exists():
            month_dir.mkdir(parents=True, exist_ok=True)
            created_dirs.add(str(month_dir.resolve()))

        report_path = month_dir / f"report_{year}_{month_name}.xlsx"
        written, duplicates = write_or_update_month_report(
            report_path=report_path,
            year=year,
            month=month,
            incoming_records=grouped_records[(year, month)],
            category_field=category_field,
        )
        stats["records_written"] += written
        stats["duplicates_skipped"] += duplicates

    stats["folders_created"] = len(created_dirs)

    app_logger.info("Processing complete.")
    app_logger.info("Files processed: %s", stats["files_processed"])
    app_logger.info("Records seen: %s", stats["records_seen"])
    app_logger.info("Records written: %s", stats["records_written"])
    app_logger.info("Duplicates skipped: %s", stats["duplicates_skipped"])
    app_logger.info("Folders created: %s", stats["folders_created"])
    app_logger.info("Invalid-date records skipped: %s", stats["invalid_date_skipped"])
    app_logger.info("Malformed files skipped: %s", stats["malformed_files"])
    app_logger.info("Errors logged at: %s", output_dir / ERROR_LOG_NAME)

    return stats


def main() -> None:
    process_json_backups(
        input_dir=INPUT_DIR,
        output_dir=OUTPUT_DIR,
        date_field=DATE_FIELD,
        category_field=CATEGORY_FIELD,
    )


if __name__ == "__main__":
    main()
